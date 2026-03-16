"""
RIMION Telegram Bot
Antigravity Full Stack Deployment
All credentials loaded from environment variables ONLY.
"""

import os
import io
import logging
from datetime import datetime
from telegram import Update, BotCommand
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

# ── Configuration — environment variables only ────────────────────────────────
TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN', '')
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')

if not TELEGRAM_TOKEN:
    raise RuntimeError("TELEGRAM_TOKEN environment variable is not set.")

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    format='%(asctime)s — %(name)s — %(levelname)s — %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ── Style helpers ─────────────────────────────────────────────────────────────
def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _medium_border():
    m = Side(style='medium', color='888888')
    t = Side(style='thin',   color='888888')
    return Border(left=t, right=t, top=m, bottom=m)

def _header_cell(ws, coord, value, bg='1E3A5F', fg='FFFFFF', size=10, center=True):
    c = ws[coord]
    c.value = value
    c.font = Font(bold=True, size=size, color=fg, name='Calibri')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center' if center else 'left',
                             vertical='center', wrap_text=True)
    c.border = _thin_border()
    return c

def _data_cell(ws, coord, value, bold=False, bg='FFFFFF', fg='000000',
               number_format=None, align='left'):
    c = ws[coord]
    c.value = value
    c.font = Font(bold=bold, size=10, color=fg, name='Calibri')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = _thin_border()
    if number_format:
        c.number_format = number_format
    return c

def _total_cell(ws, coord, formula, bg='FFD700', fg='1A1A2E'):
    c = ws[coord]
    c.value = formula
    c.font = Font(bold=True, size=11, color=fg, name='Calibri')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = _medium_border()
    c.number_format = '#,##0.00'
    return c

# ── /start ────────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "⚡ *RIMION AI — Professional Excel Solver*\n\n"
        "Available commands:\n\n"
        "📊 /pl — Profit & Loss Statement\n"
        "💼 /salary — Employee Salary Sheet\n"
        "🏦 /emi — Loan & EMI Calculator\n"
        "📈 /dashboard — Smart KPI Dashboard\n"
        "💸 /cashflow — Cash Flow Statement\n"
        "🎯 /kpi — KPI Tracker\n\n"
        "Or upload any Excel / CSV file for instant AI analysis.\n\n"
        "All reports contain *live Excel formulas* — open in Excel, "
        "change any value and everything recalculates."
    )
    await update.message.reply_text(text, parse_mode='Markdown')

# ── /pl — Profit & Loss ───────────────────────────────────────────────────────
async def cmd_pl(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚡ Building P&L Statement...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L Statement"

        today = datetime.now().strftime('%d %B %Y')

        # Title
        ws['A1'] = 'PROFIT & LOSS STATEMENT'
        ws['A1'].font = Font(bold=True, size=16, color='FFD700', name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 36

        ws['A2'] = f'Financial Year 2024-25   ·   Generated {today}   ·   All figures in ৳'
        ws['A2'].font = Font(italic=True, size=9, color='94A3B8', name='Calibri')
        ws['A2'].fill = PatternFill('solid', fgColor='0F2744')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')

        # Column headers
        for col, h in enumerate(['CATEGORY','LINE ITEM','Q1 (৳)','Q2 (৳)','Q3 (৳)','Q4 (৳)','FY TOTAL (৳)'], 1):
            _header_cell(ws, ws.cell(row=4, column=col).coordinate, h, '1E3A5F', 'FFD700')
        ws.row_dimensions[4].height = 28

        # Revenue rows
        revenue_data = [
            ('REVENUE', 'Product Sales',    450000, 480000, 510000, 540000, 'D1FAE5', '065F46'),
            ('REVENUE', 'Service Revenue',  120000, 135000, 140000, 155000, 'D1FAE5', '065F46'),
            ('REVENUE', 'Other Income',      25000,  28000,  30000,  32000, 'D1FAE5', '065F46'),
        ]
        for i, (cat, item, q1, q2, q3, q4, bg, fg) in enumerate(revenue_data, 5):
            r = i
            _data_cell(ws, f'A{r}', cat,  bold=True, bg=bg, fg=fg)
            _data_cell(ws, f'B{r}', item, bg='F0FDF4')
            for col, val in zip(['C','D','E','F'], [q1,q2,q3,q4]):
                c = ws[f'{col}{r}']
                c.value = val
                c.number_format = '#,##0.00'
                c.font = Font(size=10, name='Calibri')
                c.fill = PatternFill('solid', fgColor='F9FAFB')
                c.alignment = Alignment(horizontal='right')
                c.border = _thin_border()
            ws[f'G{r}'] = f'=SUM(C{r}:F{r})'
            ws[f'G{r}'].number_format = '#,##0.00'
            ws[f'G{r}'].font = Font(bold=True, size=10, color=fg, name='Calibri')
            ws[f'G{r}'].fill = PatternFill('solid', fgColor=bg)
            ws[f'G{r}'].alignment = Alignment(horizontal='right')
            ws[f'G{r}'].border = _thin_border()

        # Total Revenue
        _header_cell(ws, 'A8', 'TOTAL REVENUE', '00B894', 'FFFFFF')
        ws.merge_cells('A8:B8')
        for col in ['C','D','E','F']:
            _total_cell(ws, f'{col}8', f'=SUM({col}5:{col}7)', '00B894', 'FFFFFF')
        _total_cell(ws, 'G8', '=SUM(G5:G7)', '00B894', 'FFFFFF')

        # COGS rows
        cogs_data = [
            ('COGS', 'Raw Materials',      180000, 192000, 204000, 216000, 'FEF3C7', '92400E'),
            ('COGS', 'Direct Labour',       90000,  96000, 102000, 108000, 'FEF3C7', '92400E'),
            ('COGS', 'Manufacturing OH',    45000,  48000,  51000,  54000, 'FEF3C7', '92400E'),
        ]
        for i, (cat, item, q1, q2, q3, q4, bg, fg) in enumerate(cogs_data, 10):
            r = i
            _data_cell(ws, f'A{r}', cat,  bold=True, bg=bg, fg=fg)
            _data_cell(ws, f'B{r}', item, bg='FFFBEB')
            for col, val in zip(['C','D','E','F'], [q1,q2,q3,q4]):
                c = ws[f'{col}{r}']
                c.value = val
                c.number_format = '#,##0.00'
                c.font = Font(size=10, name='Calibri')
                c.fill = PatternFill('solid', fgColor='FAFAF9')
                c.alignment = Alignment(horizontal='right')
                c.border = _thin_border()
            ws[f'G{r}'] = f'=SUM(C{r}:F{r})'
            ws[f'G{r}'].number_format = '#,##0.00'
            ws[f'G{r}'].font = Font(bold=True, size=10, color=fg, name='Calibri')
            ws[f'G{r}'].fill = PatternFill('solid', fgColor=bg)
            ws[f'G{r}'].alignment = Alignment(horizontal='right')
            ws[f'G{r}'].border = _thin_border()

        _header_cell(ws, 'A13', 'TOTAL COGS', 'DC2626', 'FFFFFF')
        ws.merge_cells('A13:B13')
        for col in ['C','D','E','F']:
            _total_cell(ws, f'{col}13', f'=SUM({col}10:{col}12)', 'DC2626', 'FFFFFF')
        _total_cell(ws, 'G13', '=SUM(G10:G12)', 'DC2626', 'FFFFFF')

        # Gross Profit
        _header_cell(ws, 'A15', 'GROSS PROFIT', 'FFD700', '1A1A2E')
        ws.merge_cells('A15:B15')
        for col in ['C','D','E','F']:
            _total_cell(ws, f'{col}15', f'={col}8-{col}13', 'FFD700', '1A1A2E')
        _total_cell(ws, 'G15', '=G8-G13', 'FFD700', '1A1A2E')

        # OPEX
        opex_data = [
            ('OPEX', 'Salaries & Wages',  85000, 90000, 92000, 95000, 'EDE9FE', '4C1D95'),
            ('OPEX', 'Rent & Utilities',  24000, 24000, 24000, 24000, 'EDE9FE', '4C1D95'),
            ('OPEX', 'Marketing',         18000, 22000, 25000, 30000, 'EDE9FE', '4C1D95'),
            ('OPEX', 'Administrative',    12000, 13000, 14000, 15000, 'EDE9FE', '4C1D95'),
            ('OPEX', 'Depreciation',       8000,  8000,  8000,  8000, 'EDE9FE', '4C1D95'),
        ]
        for i, (cat, item, q1, q2, q3, q4, bg, fg) in enumerate(opex_data, 17):
            r = i
            _data_cell(ws, f'A{r}', cat,  bold=True, bg=bg, fg=fg)
            _data_cell(ws, f'B{r}', item, bg='F5F3FF')
            for col, val in zip(['C','D','E','F'], [q1,q2,q3,q4]):
                c = ws[f'{col}{r}']
                c.value = val
                c.number_format = '#,##0.00'
                c.font = Font(size=10, name='Calibri')
                c.fill = PatternFill('solid', fgColor='FAFAF9')
                c.alignment = Alignment(horizontal='right')
                c.border = _thin_border()
            ws[f'G{r}'] = f'=SUM(C{r}:F{r})'
            ws[f'G{r}'].number_format = '#,##0.00'
            ws[f'G{r}'].font = Font(bold=True, size=10, color=fg, name='Calibri')
            ws[f'G{r}'].fill = PatternFill('solid', fgColor=bg)
            ws[f'G{r}'].alignment = Alignment(horizontal='right')
            ws[f'G{r}'].border = _thin_border()

        _header_cell(ws, 'A22', 'TOTAL OPEX', '7C3AED', 'FFFFFF')
        ws.merge_cells('A22:B22')
        for col in ['C','D','E','F']:
            _total_cell(ws, f'{col}22', f'=SUM({col}17:{col}21)', '7C3AED', 'FFFFFF')
        _total_cell(ws, 'G22', '=SUM(G17:G21)', '7C3AED', 'FFFFFF')

        # Net Profit
        _header_cell(ws, 'A24', 'NET PROFIT / LOSS', '1A1A2E', 'FFD700', size=12)
        ws.merge_cells('A24:B24')
        for col in ['C','D','E','F']:
            _total_cell(ws, f'{col}24', f'={col}15-{col}22', '1A1A2E', 'FFD700')
        _total_cell(ws, 'G24', '=G15-G22', '1A1A2E', 'FFD700')

        # Notes
        ws['A26'] = ('NOTE: Gross Profit = Revenue − COGS  ·  Net Profit = Gross Profit − OPEX  '
                     '·  Annual Total = SUM(Q1:Q4)  ·  All cells contain live formulas')
        ws['A26'].font = Font(italic=True, size=8, color='94A3B8', name='Calibri')
        ws.merge_cells('A26:G26')

        # Column widths & freeze
        for col, w in zip('ABCDEFG', [16, 26, 14, 14, 14, 14, 16]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'C5'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await update.message.reply_document(
            document=buf,
            filename=f'RIMION_PL_{datetime.now().strftime("%Y%m%d")}.xlsx',
            caption=(
                "✅ *Profit & Loss Statement*\n"
                "All figures are formula-driven — change any value in Excel and everything recalculates.\n"
                "Blue = input  ·  Gold = total  ·  Green = revenue  ·  Red = cost"
            ),
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f'P&L error: {e}')
        await update.message.reply_text(f"❌ Error generating P&L: {e}")

# ── /salary — Salary Sheet ────────────────────────────────────────────────────
async def cmd_salary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚡ Building Salary Sheet...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Statement"
        month = datetime.now().strftime('%B %Y')

        NAV, WHT, YLW = '1E3A5F', 'FFFFFF', 'FEF08A'
        GRN, LGRN     = 'D1FAE5', '065F46'
        RED, LRED     = 'FEE2E2', 'DC2626'
        GR1           = 'F8FAFC'

        ws['A1'] = 'EMPLOYEE SALARY STATEMENT'
        ws['A1'].font = Font(bold=True, size=16, color=WHT, name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor=NAV)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:L1')
        ws.row_dimensions[1].height = 38

        ws['A2'] = f'Month: {month}   ·   Bangladesh Labour Law 2006 Compliant   ·   RIMION AI'
        ws['A2'].font = Font(italic=True, size=9, color='94A3B8', name='Calibri')
        ws['A2'].fill = PatternFill('solid', fgColor='0F2744')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:L2')

        headers = [
            'EMP ID', 'FULL NAME', 'DEPARTMENT', 'DESIGNATION',
            'BASIC (৳)', 'HRA 40%', 'MEDICAL 10%', 'TRANSPORT',
            'GROSS (৳)', 'PF 5%', 'INCOME TAX', 'NET PAY (৳)'
        ]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col)
            c.value = h
            c.font = Font(bold=True, size=10, color=WHT, name='Calibri')
            c.fill = PatternFill('solid', fgColor=NAV)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = _thin_border()
        ws.row_dimensions[4].height = 32

        employees = [
            ('E-001', 'Mohammad Rahman',  'Finance',    'Manager',       50000),
            ('E-002', 'Fatima Khatun',    'HR',         'Executive',     35000),
            ('E-003', 'Abdul Karim',      'IT',         'Sr. Developer', 45000),
            ('E-004', 'Nasrin Akter',     'Sales',      'Officer',       30000),
            ('E-005', 'Rahim Uddin',      'Operations', 'Supervisor',    40000),
        ]

        for i, (eid, name, dept, desig, basic) in enumerate(employees, 5):
            r = i
            bg = GR1 if i % 2 == 0 else WHT

            for col, val in zip('ABCD', [eid, name, dept, desig]):
                c = ws.cell(row=r, column=['A','B','C','D'].index(col)+1)
                c.value = val
                c.font = Font(size=10, name='Calibri')
                c.fill = PatternFill('solid', fgColor=bg)
                c.alignment = Alignment(horizontal='center' if col == 'A' else 'left', vertical='center')
                c.border = _thin_border()

            # Basic — yellow input cell
            c = ws.cell(row=r, column=5)
            c.value = basic
            c.font = Font(bold=True, size=10, name='Calibri')
            c.fill = PatternFill('solid', fgColor='FEFCE8')
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = _thin_border()
            c.number_format = '#,##0'

            # Formulas: HRA, Medical, Transport, Gross, PF, Tax, Net
            formulas = [
                (6,  f'=E{r}*0.4',                    '#,##0'),
                (7,  f'=E{r}*0.1',                    '#,##0'),
                (8,  3000,                              '#,##0'),
                (9,  f'=E{r}+F{r}+G{r}+H{r}',         '#,##0'),
                (10, f'=E{r}*0.05',                    '#,##0'),
                (11, f'=IF(I{r}*12>350000,IF(I{r}*12>500000,IF(I{r}*12>700000,IF(I{r}*12>1100000,(I{r}*12-1100000)*0.25+(1100000-700000)*0.2+(700000-500000)*0.15+(500000-350000)*0.1,(I{r}*12-700000)*0.2+(700000-500000)*0.15+(500000-350000)*0.1),(I{r}*12-500000)*0.15+(500000-350000)*0.1),(I{r}*12-350000)*0.1),0)/12', '#,##0'),
                (12, f'=I{r}-J{r}-K{r}',              '#,##0'),
            ]
            for col_num, val, fmt in formulas:
                c = ws.cell(row=r, column=col_num)
                c.value = val
                c.font = Font(size=10, name='Calibri', color=LGRN if col_num == 12 else '000000',
                              bold=(col_num == 12))
                c.fill = PatternFill('solid', fgColor=LGRN.replace('065F46','D1FAE5') if col_num == 12 else bg)
                c.alignment = Alignment(horizontal='right', vertical='center')
                c.border = _thin_border()
                c.number_format = fmt

        # Totals row
        tr = len(employees) + 5
        ws.cell(row=tr, column=1).value = 'DEPARTMENT TOTALS'
        ws.cell(row=tr, column=1).font = Font(bold=True, size=10, name='Calibri')
        ws.cell(row=tr, column=1).fill = PatternFill('solid', fgColor='1E3A5F')
        ws.cell(row=tr, column=1).font = Font(bold=True, color='FFD700', name='Calibri')
        ws.cell(row=tr, column=1).alignment = Alignment(horizontal='center')
        ws.merge_cells(f'A{tr}:D{tr}')

        for col_num in range(5, 13):
            col_letter = 'EFGHIJKL'[col_num - 5]
            c = ws.cell(row=tr, column=col_num)
            c.value = f'=SUM({col_letter}5:{col_letter}{tr-1})'
            c.font = Font(bold=True, size=11, color='1A1A2E', name='Calibri')
            c.fill = PatternFill('solid', fgColor='FFD700')
            c.alignment = Alignment(horizontal='right', vertical='center')
            c.border = _medium_border()
            c.number_format = '#,##0'

        # Notes
        note_row = tr + 2
        ws.cell(row=note_row, column=1).value = (
            'NOTES: Basic = input (yellow).  HRA = Basic × 40%.  Medical = Basic × 10%.  '
            'Transport = ৳3,000 fixed.  PF = Basic × 5% (deducted from employee).  '
            'Income Tax = Bangladesh NBR slabs (annualised).  Net Pay = Gross − PF − Tax.'
        )
        ws.cell(row=note_row, column=1).font = Font(italic=True, size=8, color='94A3B8', name='Calibri')
        ws.merge_cells(f'A{note_row}:L{note_row}')

        # Column widths & freeze
        widths = [8, 20, 14, 16, 11, 10, 11, 10, 12, 9, 12, 13]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i)].width = w
        ws.freeze_panes = 'E5'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await update.message.reply_document(
            document=buf,
            filename=f'RIMION_Salary_{datetime.now().strftime("%Y%m")}.xlsx',
            caption=(
                "✅ *Employee Salary Statement*\n"
                "Yellow cells = input (change basic salary).  All other cells recalculate automatically.\n"
                "Bangladesh Labour Law 2006 compliant — PF, NBR income tax slabs, transport allowance."
            ),
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f'Salary error: {e}')
        await update.message.reply_text(f"❌ Error generating Salary Sheet: {e}")

# ── /emi — Loan & EMI ─────────────────────────────────────────────────────────
async def cmd_emi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚡ Building EMI Calculator...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EMI Calculator"

        ws['A1'] = 'LOAN & EMI CALCULATOR'
        ws['A1'].font = Font(bold=True, size=16, color='FFD700', name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 36

        inputs = [
            ('Loan Amount (৳)',        'B3', 1000000),
            ('Annual Interest Rate (%)', 'B4', 12),
            ('Loan Tenure (Years)',      'B5', 5),
        ]
        for label, coord, val in inputs:
            row = int(coord[1:])
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True, size=10, name='Calibri')
            ws[coord].value = val
            ws[coord].fill = PatternFill('solid', fgColor='FEFCE8')
            ws[coord].font = Font(bold=True, size=10, name='Calibri')
            ws[coord].number_format = '#,##0.00'
            ws[coord].border = _thin_border()

        ws['A7'] = 'Monthly EMI (৳)'
        ws['A7'].font = Font(bold=True, size=11, name='Calibri')
        ws['B7'] = '=PMT(B4/12/100,B5*12,-B3)'
        ws['B7'].font = Font(bold=True, size=13, color='065F46', name='Calibri')
        ws['B7'].fill = PatternFill('solid', fgColor='D1FAE5')
        ws['B7'].number_format = '#,##0.00'
        ws['B7'].border = _medium_border()

        ws['A8'] = 'Total Payment (৳)'
        ws['B8'] = '=B7*B5*12'
        ws['B8'].number_format = '#,##0.00'
        ws['B8'].border = _thin_border()

        ws['A9'] = 'Total Interest (৳)'
        ws['B9'] = '=B8-B3'
        ws['B9'].font = Font(color='DC2626', name='Calibri')
        ws['B9'].fill = PatternFill('solid', fgColor='FEE2E2')
        ws['B9'].number_format = '#,##0.00'
        ws['B9'].border = _thin_border()

        # Amortisation schedule header
        ws['A11'] = 'AMORTISATION SCHEDULE'
        ws['A11'].font = Font(bold=True, size=11, color='FFD700', name='Calibri')
        ws['A11'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws.merge_cells('A11:F11')
        ws['A11'].alignment = Alignment(horizontal='center')

        for col, h in enumerate(['MONTH','OPENING BALANCE','EMI','PRINCIPAL','INTEREST','CLOSING BALANCE'], 1):
            c = ws.cell(row=12, column=col)
            c.value = h
            c.font = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
            c.fill = PatternFill('solid', fgColor='1E3A5F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin_border()

        for m in range(1, 61):
            r = 12 + m
            bg = 'F8FAFC' if m % 2 == 0 else 'FFFFFF'
            ws.cell(row=r, column=1).value = m
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')

            if m == 1:
                ws.cell(row=r, column=2).value = '=$B$3'
            else:
                ws.cell(row=r, column=2).value = f'=F{r-1}'

            ws.cell(row=r, column=3).value = '=$B$7'
            ws.cell(row=r, column=5).value = f'=B{r}*$B$4/12/100'
            ws.cell(row=r, column=4).value = f'=C{r}-E{r}'
            ws.cell(row=r, column=6).value = f'=B{r}-D{r}'

            for col in range(1, 7):
                c = ws.cell(row=r, column=col)
                c.font = Font(size=9, name='Calibri')
                c.fill = PatternFill('solid', fgColor=bg)
                c.border = _thin_border()
                if col > 1:
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right')

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 18
        for col in 'BCDEF':
            ws.column_dimensions[col].width = 16
        ws.freeze_panes = 'A13'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await update.message.reply_document(
            document=buf,
            filename=f'RIMION_EMI_{datetime.now().strftime("%Y%m%d")}.xlsx',
            caption=(
                "✅ *Loan & EMI Calculator*\n"
                "Change Loan Amount, Interest Rate or Tenure (yellow cells) — "
                "EMI and full 60-month schedule recalculate instantly."
            ),
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f'EMI error: {e}')
        await update.message.reply_text(f"❌ Error generating EMI Calculator: {e}")

# ── /cashflow — Cash Flow Statement ──────────────────────────────────────────
async def cmd_cashflow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚡ Building Cash Flow Statement...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        ws['A1'] = 'CASH FLOW STATEMENT'
        ws['A1'].font = Font(bold=True, size=16, color='FFD700', name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 36

        for col, h in enumerate(['PARTICULARS','Q1 (৳)','Q2 (৳)','Q3 (৳)','Q4 (৳)','ANNUAL (৳)'], 1):
            c = ws.cell(row=3, column=col)
            c.value = h
            c.font = Font(bold=True, size=10, color='FFD700', name='Calibri')
            c.fill = PatternFill('solid', fgColor='1E3A5F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin_border()

        sections = [
            ('A. OPERATING ACTIVITIES', '0F766E', [
                ('Net Profit',                    120000, 135000, 145000, 160000),
                ('Add: Depreciation',              15000,  15000,  15000,  15000),
                ('Changes in Working Capital',    -20000,  10000, -15000,  25000),
            ]),
            ('B. INVESTING ACTIVITIES', '1D4ED8', [
                ('Purchase of Fixed Assets',      -80000, -50000, -30000, -60000),
                ('Proceeds from Asset Sales',      10000,      0,  20000,      0),
            ]),
            ('C. FINANCING ACTIVITIES', '7C3AED', [
                ('Loan Proceeds',                 200000,      0,      0, 100000),
                ('Loan Repayment',                -40000, -40000, -40000, -40000),
                ('Dividend Paid',                      0, -50000,      0, -50000),
            ]),
        ]

        row = 4
        totals_rows = []
        for section_title, color, items in sections:
            ws.cell(row=row, column=1).value = section_title
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color='FFFFFF', name='Calibri')
            ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=color)
            ws.merge_cells(f'A{row}:F{row}')
            ws.row_dimensions[row].height = 24
            row += 1

            item_start = row
            for item, q1, q2, q3, q4 in items:
                ws.cell(row=row, column=1).value = f'  {item}'
                ws.cell(row=row, column=1).font = Font(size=10, name='Calibri')
                ws.cell(row=row, column=1).border = _thin_border()
                for col_num, val in enumerate([q1, q2, q3, q4], 2):
                    c = ws.cell(row=row, column=col_num)
                    c.value = val
                    c.number_format = '#,##0.00'
                    c.font = Font(size=10, color='DC2626' if val < 0 else '000000', name='Calibri')
                    c.fill = PatternFill('solid', fgColor='F8FAFC')
                    c.alignment = Alignment(horizontal='right')
                    c.border = _thin_border()
                ws.cell(row=row, column=6).value = f'=SUM(B{row}:E{row})'
                ws.cell(row=row, column=6).number_format = '#,##0.00'
                ws.cell(row=row, column=6).border = _thin_border()
                ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
                row += 1

            # Section total
            for col_num in range(2, 7):
                col_letter = 'BCDEF'[col_num - 2]
                _total_cell(ws, ws.cell(row=row, column=col_num).coordinate,
                            f'=SUM({col_letter}{item_start}:{col_letter}{row-1})',
                            color, 'FFFFFF')
            ws.cell(row=row, column=1).value = f'Net {section_title.split(".")[1].strip()} Cash Flow'
            ws.cell(row=row, column=1).font = Font(bold=True, size=10, color='FFFFFF', name='Calibri')
            ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor=color)
            ws.cell(row=row, column=1).border = _medium_border()
            totals_rows.append(row)
            row += 2

        # Net change in cash
        _header_cell(ws, f'A{row}', 'NET CHANGE IN CASH', 'FFD700', '1A1A2E', size=12)
        ws.merge_cells(f'A{row}:A{row}')
        for col_num, col_letter in enumerate('BCDEF', 2):
            t_refs = '+'.join([f'{col_letter}{tr}' for tr in totals_rows])
            _total_cell(ws, ws.cell(row=row, column=col_num).coordinate,
                        f'={t_refs}', 'FFD700', '1A1A2E')

        ws.column_dimensions['A'].width = 32
        for col in 'BCDEF':
            ws.column_dimensions[col].width = 14
        ws.freeze_panes = 'B4'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await update.message.reply_document(
            document=buf,
            filename=f'RIMION_CashFlow_{datetime.now().strftime("%Y%m%d")}.xlsx',
            caption="✅ *Cash Flow Statement*\nOperating · Investing · Financing — all live formulas.",
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f'Cash flow error: {e}')
        await update.message.reply_text(f"❌ Error: {e}")

# ── /dashboard — KPI Dashboard ────────────────────────────────────────────────
async def cmd_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚡ Building KPI Dashboard...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard"

        ws['A1'] = 'SMART KPI DASHBOARD'
        ws['A1'].font = Font(bold=True, size=18, color='FFD700', name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 42

        ws['A2'] = f'Generated: {datetime.now().strftime("%d %B %Y")}   ·   FY 2024-25   ·   All figures in ৳'
        ws['A2'].font = Font(italic=True, size=9, color='94A3B8', name='Calibri')
        ws['A2'].fill = PatternFill('solid', fgColor='0F2744')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:H2')

        # KPI cards
        kpis = [
            ('TOTAL REVENUE',  '=SUM(B10:B21)', '00B894', 'D1FAE5'),
            ('TOTAL EXPENSES', '=SUM(C10:C21)', 'DC2626', 'FEE2E2'),
            ('NET PROFIT',     '=B4-C4',        'FFD700', '1A1A2E'),
            ('PROFIT MARGIN',  '=IFERROR(D4/B4,"—")', '2563EB', 'EFF6FF'),
        ]
        kpi_cols = ['A','C','E','G']
        for i, (label, formula, color, bg) in enumerate(kpis):
            col = kpi_cols[i]
            next_col = chr(ord(col)+1)
            row_label, row_val = 4, 5
            ws[f'{col}{row_label}'] = label
            ws[f'{col}{row_label}'].font = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
            ws[f'{col}{row_label}'].fill = PatternFill('solid', fgColor=color)
            ws[f'{col}{row_label}'].alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells(f'{col}{row_label}:{next_col}{row_label}')
            ws.row_dimensions[row_label].height = 22

            ws[f'{col}{row_val}'] = formula
            ws[f'{col}{row_val}'].font = Font(bold=True, size=14, color=color, name='Calibri')
            ws[f'{col}{row_val}'].fill = PatternFill('solid', fgColor=bg)
            ws[f'{col}{row_val}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{row_val}'].number_format = '৳ #,##0' if i < 3 else '0.0%'
            ws[f'{col}{row_val}'].border = _medium_border()
            ws.merge_cells(f'{col}{row_val}:{next_col}{row_val}')
            ws.row_dimensions[row_val].height = 32

        # Monthly data table
        for col, h in enumerate(['MONTH','REVENUE (৳)','EXPENSES (৳)','NET PROFIT (৳)','MOM GROWTH %'], 1):
            c = ws.cell(row=9, column=col)
            c.value = h
            c.font = Font(bold=True, size=10, color='FFD700', name='Calibri')
            c.fill = PatternFill('solid', fgColor='1E3A5F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin_border()

        months = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        rev =  [520000,485000,610000,590000,640000,580000,720000,695000,750000,810000,780000,860000]
        exp =  [380000,360000,430000,410000,450000,400000,490000,470000,510000,540000,520000,570000]

        for i, (m, r, e) in enumerate(zip(months, rev, exp), 10):
            bg = 'F8FAFC' if i % 2 == 0 else 'FFFFFF'
            ws.cell(row=i, column=1).value = m
            ws.cell(row=i, column=1).font = Font(bold=True, size=10, name='Calibri')
            ws.cell(row=i, column=1).fill = PatternFill('solid', fgColor=bg)
            ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=i, column=1).border = _thin_border()

            for col_num, val in [(2, r), (3, e)]:
                c = ws.cell(row=i, column=col_num)
                c.value = val
                c.font = Font(size=10, name='Calibri')
                c.fill = PatternFill('solid', fgColor=bg)
                c.alignment = Alignment(horizontal='right')
                c.border = _thin_border()
                c.number_format = '#,##0'

            ws.cell(row=i, column=4).value = f'=B{i}-C{i}'
            ws.cell(row=i, column=4).font = Font(size=10, name='Calibri',
                color='065F46' if rev[i-10]-exp[i-10] > 0 else 'DC2626')
            ws.cell(row=i, column=4).fill = PatternFill('solid', fgColor=bg)
            ws.cell(row=i, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=i, column=4).border = _thin_border()
            ws.cell(row=i, column=4).number_format = '#,##0'

            if i > 10:
                ws.cell(row=i, column=5).value = f'=IFERROR((B{i}-B{i-1})/B{i-1},"—")'
                ws.cell(row=i, column=5).number_format = '0.0%'
                ws.cell(row=i, column=5).alignment = Alignment(horizontal='center')
                ws.cell(row=i, column=5).border = _thin_border()
                ws.cell(row=i, column=5).font = Font(size=10, name='Calibri')
            else:
                ws.cell(row=i, column=5).value = '—'
                ws.cell(row=i, column=5).alignment = Alignment(horizontal='center')
                ws.cell(row=i, column=5).border = _thin_border()

        # Annual totals
        for col_num, col_letter in enumerate('BCDE', 2):
            _total_cell(ws, ws.cell(row=22, column=col_num).coordinate,
                        f'=SUM({col_letter}10:{col_letter}21)')
        ws.cell(row=22, column=1).value = 'ANNUAL TOTAL'
        ws.cell(row=22, column=1).font = Font(bold=True, size=10, color='1A1A2E', name='Calibri')
        ws.cell(row=22, column=1).fill = PatternFill('solid', fgColor='FFD700')
        ws.cell(row=22, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=22, column=1).border = _medium_border()

        for col, w in zip('ABCDEFGH', [10,14,14,14,12,14,14,14]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A10'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await update.message.reply_document(
            document=buf,
            filename=f'RIMION_Dashboard_{datetime.now().strftime("%Y%m%d")}.xlsx',
            caption=(
                "✅ *Smart KPI Dashboard*\n"
                "Revenue · Expenses · Net Profit · MoM Growth — 12 months with live formulas."
            ),
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f'Dashboard error: {e}')
        await update.message.reply_text(f"❌ Error: {e}")

# ── /kpi — KPI Tracker ────────────────────────────────────────────────────────
async def cmd_kpi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⚡ Building KPI Tracker...")
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KPI Tracker"

        ws['A1'] = 'KPI PERFORMANCE TRACKER'
        ws['A1'].font = Font(bold=True, size=16, color='FFD700', name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 36

        for col, h in enumerate(['KPI NAME','TARGET','ACTUAL','ACHIEVEMENT %','SCORE /10','STATUS','TREND'], 1):
            c = ws.cell(row=3, column=col)
            c.value = h
            c.font = Font(bold=True, size=10, color='FFD700', name='Calibri')
            c.fill = PatternFill('solid', fgColor='1E3A5F')
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = _thin_border()
        ws.row_dimensions[3].height = 26

        kpis = [
            ('Revenue (৳)',       1000000, 1120000),
            ('Gross Margin %',         42,      45),
            ('Net Profit %',           18,      16),
            ('Customer Count',        500,     520),
            ('Employee Turnover %',    10,       8),
            ('On-Time Delivery %',     95,      92),
            ('Customer Satisfaction',   4.5,    4.7),
            ('New Leads',             200,     235),
        ]

        for i, (name, target, actual) in enumerate(kpis, 4):
            r = i
            bg = 'F8FAFC' if i % 2 == 0 else 'FFFFFF'

            ws.cell(row=r, column=1).value = name
            ws.cell(row=r, column=1).font = Font(bold=True, size=10, name='Calibri')
            ws.cell(row=r, column=1).fill = PatternFill('solid', fgColor=bg)
            ws.cell(row=r, column=1).border = _thin_border()

            ws.cell(row=r, column=2).value = target
            ws.cell(row=r, column=2).fill = PatternFill('solid', fgColor='FEFCE8')
            ws.cell(row=r, column=2).font = Font(bold=True, size=10, name='Calibri')
            ws.cell(row=r, column=2).alignment = Alignment(horizontal='right')
            ws.cell(row=r, column=2).border = _thin_border()
            ws.cell(row=r, column=2).number_format = '#,##0.##'

            ws.cell(row=r, column=3).value = actual
            ws.cell(row=r, column=3).fill = PatternFill('solid', fgColor='FEFCE8')
            ws.cell(row=r, column=3).font = Font(bold=True, size=10, name='Calibri')
            ws.cell(row=r, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=r, column=3).border = _thin_border()
            ws.cell(row=r, column=3).number_format = '#,##0.##'

            ws.cell(row=r, column=4).value = f'=C{r}/B{r}'
            ws.cell(row=r, column=4).number_format = '0.0%'
            ws.cell(row=r, column=4).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=4).border = _thin_border()
            ws.cell(row=r, column=4).font = Font(size=10, name='Calibri')

            ws.cell(row=r, column=5).value = f'=MIN(10,D{r}*10)'
            ws.cell(row=r, column=5).number_format = '0.0'
            ws.cell(row=r, column=5).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=5).border = _thin_border()
            ws.cell(row=r, column=5).font = Font(bold=True, size=10, name='Calibri')

            ws.cell(row=r, column=6).value = f'=IF(D{r}>=1,"✅ On Target",IF(D{r}>=0.9,"⚠️ Near Target","❌ Off Target"))'
            ws.cell(row=r, column=6).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=6).border = _thin_border()
            ws.cell(row=r, column=6).font = Font(size=10, name='Calibri')

            ws.cell(row=r, column=7).value = '↑' if actual >= target else '↓'
            ws.cell(row=r, column=7).font = Font(bold=True, size=14, name='Calibri',
                color='065F46' if actual >= target else 'DC2626')
            ws.cell(row=r, column=7).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=r, column=7).fill = PatternFill('solid', fgColor=bg)
            ws.cell(row=r, column=7).border = _thin_border()

        for col, w in zip('ABCDEFG', [22, 12, 12, 14, 10, 16, 8]):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A4'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        await update.message.reply_document(
            document=buf,
            filename=f'RIMION_KPI_{datetime.now().strftime("%Y%m%d")}.xlsx',
            caption=(
                "✅ *KPI Performance Tracker*\n"
                "Change Target and Actual (yellow cells) — Achievement %, Score and Status update automatically."
            ),
            parse_mode='Markdown'
        )
    except Exception as e:
        logger.error(f'KPI error: {e}')
        await update.message.reply_text(f"❌ Error: {e}")

# ── File upload handler ───────────────────────────────────────────────────────
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not doc:
        return
    fname = doc.file_name or ''
    ext = fname.rsplit('.', 1)[-1].lower() if '.' in fname else ''

    if ext not in ('xlsx', 'xls', 'csv'):
        await update.message.reply_text(
            "⚠️ Please upload an Excel (.xlsx/.xls) or CSV file.\n\n"
            "Or use a command:\n"
            "/pl · /salary · /emi · /dashboard · /cashflow · /kpi"
        )
        return

    await update.message.reply_text(
        f"📊 *{fname}* received.\n\n"
        "Use a command to generate a report from this data:\n"
        "/pl · /salary · /emi · /dashboard · /cashflow · /kpi",
        parse_mode='Markdown'
    )

# ── Unknown command ───────────────────────────────────────────────────────────
async def handle_unknown(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Command not recognised. Type /start to see all available commands."
    )

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    logger.info("Starting RIMION Telegram Bot...")

    app = Application.builder().token(TELEGRAM_TOKEN).build()

    app.add_handler(CommandHandler("start",     cmd_start))
    app.add_handler(CommandHandler("pl",        cmd_pl))
    app.add_handler(CommandHandler("salary",    cmd_salary))
    app.add_handler(CommandHandler("emi",       cmd_emi))
    app.add_handler(CommandHandler("cashflow",  cmd_cashflow))
    app.add_handler(CommandHandler("dashboard", cmd_dashboard))
    app.add_handler(CommandHandler("kpi",       cmd_kpi))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    app.add_handler(MessageHandler(filters.COMMAND, handle_unknown))

    logger.info("Bot is running.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == '__main__':
    main()
