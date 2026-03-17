"""
RIMION Pro Backend API
Empowering Professional Excel Productivity with AI.
"""

import os
import io
import json
import base64
import logging
from typing import List, Optional
from datetime import datetime

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
import google.generativeai as genai
from dotenv import load_dotenv

load_dotenv()

from utils_security import get_gemini_key, get_telegram_token

# ── Configuration ─────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

GEMINI_API_KEY = get_gemini_key()
if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

app = FastAPI(title="RIMION API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _header_cell(ws, coord, value, bg='1E3A5F', fg='FFFFFF'):
    c = ws[coord]
    c.value = value
    c.font = Font(bold=True, color=fg, name='Calibri')
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = _thin_border()

# ── Endpoints ─────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root():
    try:
        with open("RIMION_LIVE.html", "r", encoding="utf-8") as f:
            content = f.read()
            # Inject dynamic backend URL if needed or just serve as is
            return content
    except FileNotFoundError:
        return HTMLResponse("<h2>RIMION API running. Upload RIMION_LIVE.html.</h2>")

@app.get("/health")
async def health():
    return {"status": "ok", "service": "RIMION Pro API", "ai_ready": bool(GEMINI_API_KEY)}

# ── AI Chat ───────────────────────────────────────────────────────────────────

@app.post("/chat")
async def chat_with_file(
    file: UploadFile = File(...),
    message: str = Form(...),
    history: str = Form("[]")
):
    if not GEMINI_API_KEY:
        return JSONResponse({"reply": "API Key missing on server. Please use client-side AI or set GEMINI_API_KEY environment variable."}, status_code=500)

    try:
        contents = await file.read()
        df = None
        if file.filename.endswith('.csv'):
            df = pd.read_csv(io.BytesIO(contents))
        else:
            df = pd.read_excel(io.BytesIO(contents))

        # Context construction - ENHANCED
        context = f"User has uploaded a file: {file.filename}\n"
        context += f"Shape: {df.shape[0]} rows, {df.shape[1]} columns\n"
        context += f"Columns: {', '.join(df.columns.tolist())}\n"
        
        # Add descriptive statistics for numeric columns
        num_cols = df.select_dtypes(include=['number']).columns.tolist()
        if num_cols:
            context += f"\nKey Statistics for numeric columns ({', '.join(num_cols)}):\n"
            stats = df[num_cols].describe().to_string()
            context += f"{stats}\n"
            
        # Add sample data (top and bottom to see range)
        context += f"\nSample Data (First 3 rows):\n{df.head(3).to_string()}\n"
        context += f"\nSample Data (Last 2 rows):\n{df.tail(2).to_string()}\n"

        model = genai.GenerativeModel("gemini-1.5-flash")
        
        prompt = (
            "You are RIMION AI, the world's most advanced Excel and Financial analyst. "
            "You have full access to the user's data context provided below. "
            "Help the user with their specific request: questions, formatting instructions, or formula generation.\n\n"
            f"FILE CONTEXT:\n{context}\n\n"
            f"USER MESSAGE: {message}\n\n"
            "INSTRUCTIONS:\n"
            "1. Be precise and data-driven.\n"
            "2. If suggesting formulas, use valid Excel syntax wrapped in backticks.\n"
            "3. If the user asks for insights, look at the statistics provided.\n"
            "4. Maintain a professional, executive tone."
        )

        response = model.generate_content(prompt)
        reply = response.text

        return JSONResponse({
            "status": "success",
            "reply": reply,
            "has_modified_file": False 
        })
    except Exception as e:
        logger.error(f"Chat error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)

# ── Audit ─────────────────────────────────────────────────────────────────────

@app.post("/audit")
@app.post("/api/audit")
async def audit_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        # Openpyxl for structural audit
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=False)
        ws = wb.active
        
        issues = []
        
        # 1. Structural Scan (Limited for performance)
        max_scan_row = min(ws.max_row, 500)
        max_scan_col = min(ws.max_column, 26)
        
        for r in range(1, max_scan_row + 1):
            for c in range(1, max_scan_col + 1):
                cell = ws.cell(row=r, column=c)
                
                # Formula Errors
                if isinstance(cell.value, str) and cell.value.startswith('#'):
                    issues.append({"cell": cell.coordinate, "issue": f"Formula Error: {cell.value}", "severity": "high", "fix": "Check formula references"})
                
                # Numbers as text
                if isinstance(cell.value, str) and not cell.value.startswith('='):
                    clean_val = cell.value.replace(',', '').replace('৳', '').replace('$', '').strip()
                    if clean_val.replace('.', '', 1).isdigit():
                        issues.append({"cell": cell.coordinate, "issue": "Number stored as text", "severity": "medium", "fix": "Convert to Number format"})

        # 2. Data Health Scan (Pandas)
        df = None
        try:
            df = pd.read_excel(io.BytesIO(contents))
            # Duplicates
            duplicates = df.duplicated()
            dup_count = duplicates.sum()
            if dup_count > 0:
                issues.append({"issue": f"Found {dup_count} exact duplicate rows", "severity": "high", "fix": "Remove duplicates via Data Tab"})
            
            # Missing Value analysis
            missing = df.isnull().sum()
            for col, count in missing.items():
                if count > 0:
                    pct = (count / len(df)) * 100
                    severity = "high" if pct > 30 else "medium"
                    issues.append({"issue": f"Column '{col}' is missing {count} values ({pct:.1f}%)", "severity": severity, "fix": "Clean data or fill blanks"})
        except Exception as e:
            logger.warning(f"Pandas audit skip: {e}")

        # AI Summary Generation
        summary_prompt = (
            "You are an Excel Auditor. Summarize the following audit findings into a 2-sentence professional executive summary.\n"
            f"Findings: {len(issues)} issues total. Data type: {df.dtypes.to_dict() if df is not None else 'Unknown'}.\n"
            "Focus on data integrity and professional readiness."
        )
        
        ai_summary = "Audit complete. No critical structural issues found."
        try:
            if issues and GEMINI_API_KEY:
                model = genai.GenerativeModel("gemini-1.5-flash")
                response = model.generate_content(summary_prompt)
                ai_summary = response.text
        except:
            pass

        return JSONResponse({
            "status": "success",
            "stats": {
                "issues_found": len(issues),
                "issues_fixed": 0,
                "rows": ws.max_row,
                "cols": ws.max_column
            },
            "issues": issues[:30],
            "ai_summary": ai_summary,
            "has_fixes": False 
        })
    except Exception as e:
        logger.error(f"Audit error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)

# ── Modify ────────────────────────────────────────────────────────────────────

@app.post("/modify")
async def modify_excel(
    file: UploadFile = File(...),
    option_id: str = Form(...),
    params: str = Form("{}")
):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        # Apply Logic Based on option_id
        if option_id == 'fmt_professional':
            # Header Row Styling
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill("solid", fgColor="1E3A5F")
                cell.alignment = Alignment(horizontal="center")
            # Zebra Rows
            for r in range(2, ws.max_row + 1):
                if r % 2 == 0:
                    for cell in ws[r]:
                        cell.fill = PatternFill("solid", fgColor="F1F5F9")
            # Auto-fit (rough approximation)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 15

        elif option_id == 'cln_trim':
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = cell.value.strip()

        elif option_id == 'fmt_freeze':
            ws.freeze_panes = 'A2'

        # More options can be implemented here...

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="RIMION_MOD_{file.filename}"'}
        )
    except Exception as e:
        logger.error(f"Modify error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)

# ── Report Generation ─────────────────────────────────────────────────────────

@app.post("/api/generate/{report_type}")
@app.post("/generate/{report_type}")
async def generate_report(report_type: str):
    try:
        # Use the Telegram Bot's logic approach for high quality
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.upper()
        
        today = datetime.now().strftime('%d %B %Y')
        
        # Title
        ws['A1'] = f'RIMION — {report_type.upper()} REPORT'
        ws['A1'].font = Font(bold=True, size=16, color='FFD700', name='Calibri')
        ws['A1'].fill = PatternFill('solid', fgColor='1A1A2E')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:G1')
        ws.row_dimensions[1].height = 36
        
        ws['A2'] = f'Generated: {today}   ·   Professional Edition   ·   All figures in ৳'
        ws['A2'].font = Font(italic=True, size=9, color='94A3B8', name='Calibri')
        ws['A2'].fill = PatternFill('solid', fgColor='0F2744')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:G2')

        # Dummy professional data based on type
        if report_type == 'pl':
            headers = ['CATEGORY','LINE ITEM','Q1 (৳)','Q2 (৳)','Q3 (৳)','Q4 (৳)','FY TOTAL (৳)']
            for col, h in enumerate(headers, 1):
                _header_cell(ws, ws.cell(row=4, column=col).coordinate, h)
            # Add some demo formula rows...
            ws['A5'] = 'REVENUE'; ws['B5'] = 'Sales'; ws['C5'] = 500000; ws['D5'] = 550000; ws['E5'] = 600000; ws['F5'] = 650000; ws['G5'] = '=SUM(C5:F5)'
            ws['G5'].number_format = '#,##0.00'
        
        elif report_type == 'salary':
            headers = ['EMP ID', 'NAME', 'DEPT', 'BASIC', 'HRA', 'PF', 'TAX', 'NET']
            for col, h in enumerate(headers, 1):
                _header_cell(ws, ws.cell(row=4, column=col).coordinate, h)
            ws['A5'] = 'E001'; ws['B5'] = 'John Doe'; ws['C5'] = 'Finance'; ws['D5'] = 50000; ws['E5'] = '=D5*0.4'; ws['F5'] = '=D5*0.05'; ws['G5'] = 0; ws['H5'] = '=D5+E5-F5-G5'

        # Auto width
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        for c in 'CDEFG':
            ws.column_dimensions[c].width = 14

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="RIMION_{report_type}.xlsx"'}
        )
    except Exception as e:
        logger.error(f"Generate error: {e}")
        return JSONResponse({"status": "error", "message": str(e)}, status_code=500)

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    # Note: On Render, PORT is provided as an env var
    uvicorn.run(app, host="0.0.0.0", port=port)
